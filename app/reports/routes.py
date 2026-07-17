from fastapi.responses import Response

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Table
from reportlab.platypus import TableStyle
from reportlab.platypus import Paragraph
from reportlab.platypus import Spacer
from app.suppliers.models import Supplier
from reportlab.lib.styles import getSampleStyleSheet

from io import BytesIO
from fastapi import APIRouter
from fastapi import Depends
from fastapi import Request
from fastapi.responses import HTMLResponse
from fastapi.responses import RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from fastapi.responses import StreamingResponse

from app.config.database import get_db

from app.billing.service import BillingService
from io import BytesIO

from fastapi.responses import Response

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from app.purchase.service import PurchaseService
from app.inventory.service import InventoryService
from app.customers.models import Customer

from datetime import datetime

from app.auth.dependencies import login_required

router = APIRouter(
    dependencies=[Depends(login_required)]
)

templates = Jinja2Templates(
    directory="app/templates"
)

# --------------------------------------------------
# Sales Report
# --------------------------------------------------
def add_page_number(canvas, doc):

    canvas.saveState()

    canvas.setFont(
        "DejaVuSans",
        9
    )

    canvas.drawRightString(
        800,
        20,
        f"Page {canvas.getPageNumber()}"
    )

    canvas.restoreState()
@router.get(
    "/reports/sales",
    response_class=HTMLResponse
)
async def sales_report(
    request: Request,
    customer_id: str = "",
    from_date: str = "",
    to_date: str = "",
    user=Depends(login_required),
    db: Session = Depends(get_db),
):

    if isinstance(user, RedirectResponse):
        return user

    invoices = BillingService.get_all(db)

    # -------------------------------
    # Date Filters
    # -------------------------------

    if from_date:

        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            invoice
            for invoice in invoices
            if invoice.invoice_date >= from_dt
        ]

    if to_date:

        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            invoice
            for invoice in invoices
            if invoice.invoice_date <= to_dt
        ]

    # -------------------------------
    # Customer Filter
    # -------------------------------

    selected_customer_id = None

    if customer_id and customer_id.isdigit():

        selected_customer_id = int(customer_id)

        invoices = [
            invoice
            for invoice in invoices
            if invoice.customer_id == selected_customer_id
        ]

    # -------------------------------
    # Summary
    # -------------------------------

    total_sales = sum(
        float(invoice.grand_total or 0)
        for invoice in invoices
    )

    total_invoices = len(invoices)

    avg_invoice = (
        total_sales / total_invoices
        if total_invoices else 0
    )

    customers = (
        db.query(Customer)
        .order_by(Customer.customer_name)
        .all()
    )

    return templates.TemplateResponse(
        request=request,
        name="reports/sales.html",
        context={
            "invoices": invoices,
            "customers": customers,
            "customer_id": selected_customer_id,
            "from_date": from_date,
            "to_date": to_date,
            "total_sales": total_sales,
            "total_invoices": total_invoices,
            "avg_invoice": avg_invoice,
        },
    )


# --------------------------------------------------
# Sales Report PDF
# --------------------------------------------------

@router.get("/reports/sales/pdf")
async def sales_report_pdf(
    customer_id: str = "",
    from_date: str = "",
    to_date: str = "",
    db: Session = Depends(get_db),
):

    pdfmetrics.registerFont(
        TTFont(
            "DejaVuSans",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        )
    )

    invoices = BillingService.get_all(db)

    # -------------------------
    # Date Filters
    # -------------------------

    if from_date:

        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            x for x in invoices
            if x.invoice_date >= from_dt
        ]

    if to_date:

        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            x for x in invoices
            if x.invoice_date <= to_dt
        ]

    # -------------------------
    # Customer Filter
    # -------------------------

    customer_name = "All Customers"

    if customer_id and customer_id.isdigit():

        customer_id_int = int(customer_id)

        invoices = [
            x for x in invoices
            if x.customer_id == customer_id_int
        ]

        customer = (
            db.query(Customer)
            .filter(Customer.id == customer_id_int)
            .first()
        )

        if customer:
            customer_name = customer.customer_name

    total_sales = sum(
        float(x.grand_total or 0)
        for x in invoices
    )

    # -------------------------
    # PDF
    # -------------------------

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    # -------------------------
    # Logo
    # -------------------------

    try:

        logo = Image(
            "app/static/uploads/logo/logo.png",
            width=60,
            height=60
        )

        elements.append(logo)

    except Exception:
        pass

    # -------------------------
    # Company Header
    # -------------------------

    elements.append(
        Paragraph(
            "<b>ADS ERP</b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Sales Report",
            styles["Title"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    # -------------------------
    # Filters Table
    # -------------------------

    filter_data = [
        ["From Date", from_date or "-"],
        ["To Date", to_date or "-"],
        ["Customer", customer_name],
        [
            "Generated On",
            datetime.now().strftime(
                "%d-%m-%Y %I:%M %p"
            ),
        ],
    ]

    filter_table = Table(
        filter_data,
        colWidths=[120, 300]
    )

    filter_table.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
            ("BACKGROUND", (0, 0), (0, -1),
             colors.lightgrey),
        ])
    )

    elements.append(filter_table)

    elements.append(
        Spacer(1, 20)
    )

    # -------------------------
    # Sales Table
    # -------------------------

    data = [[
        "Invoice No",
        "Date",
        "Customer",
        "Status",
        "Amount (₹)"
    ]]

    for invoice in invoices:

        data.append([
            invoice.invoice_no,
            invoice.invoice_date.strftime(
                "%d-%m-%Y"
            ),
            (
                invoice.customer.customer_name
                if invoice.customer
                else "-"
            ),
            invoice.payment_status,
            f"Rs. {float(invoice.grand_total):,.2f}",
        ])

    data.append([
        "",
        "",
        "",
        "GRAND TOTAL",
        f"Rs. {total_sales:,.2f}"
    ])

    table = Table(
        data,
        colWidths=[
            120,
            120,
            250,
            120,
            120,
        ]
    )

    table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#0d3b66")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "DejaVuSans"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.black
            ),

            (
                "BACKGROUND",
                (0, -1),
                (-1, -1),
                colors.HexColor("#f2f2f2")
            ),

            (
                "FONTNAME",
                (0, -1),
                (-1, -1),
                "DejaVuSans"
            ),

        ])
    )

    elements.append(table)

    elements.append(
        Spacer(1, 20)
    )

    # -------------------------
    # Footer
    # -------------------------

    elements.append(
        Paragraph(
            f"<b>Total Sales :</b> Rs. {total_sales:,.2f}",
            styles["Heading3"]
        )
    )

    elements.append(
        Paragraph(
            "Generated by ADS ERP",
            styles["Normal"]
        )
    )

    doc.build(
        elements,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    pdf = buffer.getvalue()

    buffer.close()

    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "inline; filename=sales_report.pdf"
        },
    )

@router.get("/reports/sales/excel")
async def sales_report_excel(
    customer_id: str = "",
    from_date: str = "",
    to_date: str = "",
    db: Session = Depends(get_db),
):

    invoices = BillingService.get_all(db)

    if from_date:
        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            x for x in invoices
            if x.invoice_date >= from_dt
        ]

    if to_date:
        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        invoices = [
            x for x in invoices
            if x.invoice_date <= to_dt
        ]

    if customer_id:
        invoices = [
            x for x in invoices
            if str(x.customer_id) == customer_id
        ]

    total_sales = sum(
        float(x.grand_total or 0)
        for x in invoices
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Sales Report"

    ws.merge_cells("A1:F1")

    ws["A1"] = "ADS ERP - Sales Report"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Invoice No",
        "Date",
        "Customer",
        "Payment Status",
        "Payment Mode",
        "Amount"
    ]

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=3,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

    row_num = 4

    for invoice in invoices:

        ws.cell(
            row=row_num,
            column=1,
            value=invoice.invoice_no
        )

        ws.cell(
            row=row_num,
            column=2,
            value=str(invoice.invoice_date)
        )

        ws.cell(
            row=row_num,
            column=3,
            value=(
                invoice.customer.customer_name
                if invoice.customer
                else "-"
            )
        )

        ws.cell(
            row=row_num,
            column=4,
            value=invoice.payment_status
        )

        ws.cell(
            row=row_num,
            column=5,
            value=invoice.payment_mode
        )

        ws.cell(
            row=row_num,
            column=6,
            value=float(invoice.grand_total)
        )

        row_num += 1

    ws.cell(
        row=row_num,
        column=5,
        value="GRAND TOTAL"
    ).font = Font(
        bold=True
    )

    ws.cell(
        row=row_num,
        column=6,
        value=total_sales
    ).font = Font(
        bold=True
    )

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=sales_report.xlsx"
        }
    )

@router.get("/reports/purchase/pdf")
async def purchase_report_pdf(
    supplier_id: str = "",
    from_date: str = "",
    to_date: str = "",
    db: Session = Depends(get_db),
):

    from app.purchase.models import Purchase

    purchases = PurchaseService.get_all(db)

    if from_date:
        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date >= from_dt
        ]

    if to_date:
        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date <= to_dt
        ]

    if supplier_id:
        purchases = [
            p for p in purchases
            if str(p.supplier_id) == supplier_id
        ]

    total_purchase = sum(
        float(p.grand_total or 0)
        for p in purchases
    )

    supplier_name = "All Suppliers"

    if supplier_id:

        supplier = (
            db.query(Supplier)
            .filter(Supplier.id == int(supplier_id))
            .first()
        )

        if supplier:
            supplier_name = supplier.supplier_name

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()

    elements = []

    logo_path = "app/static/uploads/logo/logo.png"

    try:

        logo = Image(
            logo_path,
            width=50,
            height=50
        )

        logo.hAlign = "CENTER"

        elements.append(logo)

    except:
        pass

    title = Paragraph(
        "<b>ADS ERP</b>",
        styles["Title"]
    )

    elements.append(title)

    elements.append(
        Spacer(1, 10)
    )

    report_title = Paragraph(
        "<b>Purchase Report</b>",
        styles["Heading1"]
    )

    elements.append(report_title)

    elements.append(
        Spacer(1, 10)
    )

    generated_on = datetime.now().strftime(
        "%d-%m-%Y %I:%M %p"
    )

    filter_data = [
        ["From Date", from_date or "-"],
        ["To Date", to_date or "-"],
        ["Supplier", supplier_name],
        ["Generated On", generated_on],
    ]

    filter_table = Table(
        filter_data,
        colWidths=[140, 350]
    )

    filter_table.setStyle(
        TableStyle([
            ("GRID", (0,0), (-1,-1), 1, colors.black),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ])
    )

    elements.append(filter_table)

    elements.append(
        Spacer(1, 20)
    )

    table_data = [[
        "Purchase No",
        "Date",
        "Supplier",
        "Payment Mode",
        "Amount (₹)"
    ]]

    for purchase in purchases:

        table_data.append([
            purchase.purchase_no,
            purchase.purchase_date.strftime("%d-%m-%Y"),
            purchase.supplier.supplier_name
            if purchase.supplier else "-",
            purchase.payment_mode or "-",
            f"₹ {float(purchase.grand_total):,.2f}"
        ])

    table_data.append([
        "",
        "",
        "",
        "GRAND TOTAL",
        f"₹ {total_purchase:,.2f}"
    ])

    report_table = Table(
        table_data,
        colWidths=[120, 120, 250, 120, 120]
    )

    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#15406A")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("GRID", (0,0), (-1,-1), 1, colors.black),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME", (3,-1), (4,-1), "Helvetica-Bold"),
        ])
    )

    elements.append(report_table)

    elements.append(
        Spacer(1, 20)
    )

    elements.append(
        Paragraph(
            f"<b>Total Purchase : ₹ {total_purchase:,.2f}</b>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            "Generated by ADS ERP",
            styles["Normal"]
        )
    )

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "inline; filename=purchase_report.pdf"
        },
    )

@router.get("/reports/purchase/excel")
async def purchase_report_excel(
    supplier_id: str = "",
    from_date: str = "",
    to_date: str = "",
    db: Session = Depends(get_db),
):

    purchases = PurchaseService.get_all(db)

    if from_date:
        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date >= from_dt
        ]

    if to_date:
        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date <= to_dt
        ]

    if supplier_id:
        purchases = [
            p for p in purchases
            if str(p.supplier_id) == supplier_id
        ]

    total_purchase = sum(
        float(p.grand_total or 0)
        for p in purchases
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Purchase Report"

    ws.merge_cells("A1:E1")

    ws["A1"] = "ADS ERP - Purchase Report"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Purchase No",
        "Date",
        "Supplier",
        "Payment Mode",
        "Amount"
    ]

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=3,
            column=col_num
        )

        cell.value = header

        cell.font = Font(bold=True)

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

    row_num = 4

    for purchase in purchases:

        ws.cell(
            row=row_num,
            column=1,
            value=purchase.purchase_no
        )

        ws.cell(
            row=row_num,
            column=2,
            value=str(purchase.purchase_date)
        )

        ws.cell(
            row=row_num,
            column=3,
            value=purchase.supplier.supplier_name
            if purchase.supplier else "-"
        )

        ws.cell(
            row=row_num,
            column=4,
            value=purchase.payment_mode
        )

        ws.cell(
            row=row_num,
            column=5,
            value=float(purchase.grand_total)
        )

        row_num += 1

    ws.cell(
        row=row_num,
        column=4,
        value="GRAND TOTAL"
    ).font = Font(bold=True)

    ws.cell(
        row=row_num,
        column=5,
        value=total_purchase
    ).font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=purchase_report.xlsx"
        }
    )
# --------------------------------------------------
# Purchase Report
# --------------------------------------------------

@router.get(
    "/reports/purchase",
    response_class=HTMLResponse
)
async def purchase_report(
    request: Request,
    supplier_id: str = "",
    from_date: str = "",
    to_date: str = "",
    db: Session = Depends(get_db),
):

    from app.purchase.models import Purchase
    from app.suppliers.models import Supplier

    purchases = PurchaseService.get_all(db)

    if from_date:
        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date >= from_dt
        ]

    if to_date:
        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d"
        ).date()

        purchases = [
            p for p in purchases
            if p.purchase_date <= to_dt
        ]

    if supplier_id:
        purchases = [
            p for p in purchases
            if str(p.supplier_id) == supplier_id
        ]

    total_purchase = sum(
        float(p.grand_total or 0)
        for p in purchases
    )

    total_bills = len(purchases)

    avg_purchase = (
        total_purchase / total_bills
        if total_bills else 0
    )

    suppliers = (
        db.query(Supplier)
        .order_by(Supplier.supplier_name)
        .all()
    )

    return templates.TemplateResponse(
        request=request,
        name="reports/purchase.html",
        context={
            "purchases": purchases,
            "suppliers": suppliers,
            "supplier_id": supplier_id,
            "from_date": from_date,
            "to_date": to_date,
            "total_purchase": total_purchase,
            "total_bills": total_bills,
            "avg_purchase": avg_purchase,
        }
    )


# --------------------------------------------------
# Inventory Report
# --------------------------------------------------

@router.get(
    "/reports/inventory",
    response_class=HTMLResponse
)
async def inventory_report(
    request: Request,
    stock_status: str = "",
    search: str = "",
    db: Session = Depends(get_db),
):

    products = InventoryService.get_current_stock(db)

    if search:

        products = [
            p for p in products
            if search.lower()
            in p.product_name.lower()
        ]

    if stock_status == "low":

        products = [
            p for p in products
            if float(p.current_stock)
            <= float(p.minimum_stock)
        ]

    elif stock_status == "instock":

        products = [
            p for p in products
            if float(p.current_stock)
            > float(p.minimum_stock)
        ]

    total_products = len(products)

    total_qty = sum(
        float(p.current_stock or 0)
        for p in products
    )

    low_stock_count = len(
        [
            p for p in products
            if float(p.current_stock)
            <= float(p.minimum_stock)
        ]
    )

    inventory_value = sum(
        float(p.current_stock or 0)
        * float(p.purchase_price or 0)
        for p in products
    )

    return templates.TemplateResponse(
        request=request,
        name="reports/inventory.html",
        context={
            "products": products,
            "stock_status": stock_status,
            "search": search,
            "total_products": total_products,
            "total_qty": total_qty,
            "low_stock_count": low_stock_count,
            "inventory_value": inventory_value,
        },
    )

@router.get("/reports/inventory/pdf")
async def inventory_report_pdf(
    stock_status: str = "",
    search: str = "",
    db: Session = Depends(get_db),
):

    products = InventoryService.get_current_stock(db)

    if search:
        products = [
            p for p in products
            if search.lower() in p.product_name.lower()
        ]

    if stock_status == "low":
        products = [
            p for p in products
            if float(p.current_stock)
            <= float(p.minimum_stock)
        ]

    elif stock_status == "instock":
        products = [
            p for p in products
            if float(p.current_stock)
            > float(p.minimum_stock)
        ]

    inventory_value = sum(
        float(p.current_stock or 0)
        * float(p.purchase_price or 0)
        for p in products
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()

    elements = []

    try:

        logo = Image(
            "app/static/uploads/logo/logo.png",
            width=50,
            height=50,
        )

        logo.hAlign = "CENTER"

        elements.append(logo)

    except:
        pass

    elements.append(
        Paragraph(
            "<b>ADS ERP</b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<b>Inventory Report</b>",
            styles["Heading1"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(
        Paragraph(
            f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    data = [[
        "Product",
        "Purchase",
        "Selling",
        "Current",
        "Minimum",
        "Value",
        "Status"
    ]]

    for p in products:

        value = (
            float(p.current_stock)
            * float(p.purchase_price)
        )

        status = (
            "Low Stock"
            if float(p.current_stock)
            <= float(p.minimum_stock)
            else "In Stock"
        )

        data.append([
            p.product_name,
            f"{float(p.purchase_price):,.2f}",
            f"{float(p.selling_price):,.2f}",
            str(p.current_stock),
            str(p.minimum_stock),
            f"{value:,.2f}",
            status,
        ])

    table = Table(
        data,
        colWidths=[180,70,70,70,70,100,90]
    )

    table.setStyle(
        TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#15406A")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),1,colors.black),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ])
    )

    elements.append(table)

    elements.append(
        Spacer(1, 15)
    )

    elements.append(
        Paragraph(
            f"<b>Total Inventory Value : Rs {inventory_value:,.2f}</b>",
            styles["Heading2"]
        )
    )

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "inline; filename=inventory_report.pdf"
        }
    )

@router.get("/reports/inventory/excel")
async def inventory_report_excel(
    stock_status: str = "",
    search: str = "",
    db: Session = Depends(get_db),
):

    products = InventoryService.get_current_stock(db)

    if search:
        products = [
            p for p in products
            if search.lower() in p.product_name.lower()
        ]

    if stock_status == "low":
        products = [
            p for p in products
            if float(p.current_stock)
            <= float(p.minimum_stock)
        ]

    elif stock_status == "instock":
        products = [
            p for p in products
            if float(p.current_stock)
            > float(p.minimum_stock)
        ]

    inventory_value = sum(
        float(p.current_stock or 0)
        * float(p.purchase_price or 0)
        for p in products
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventory Report"

    ws.merge_cells("A1:G1")

    ws["A1"] = "ADS ERP - Inventory Report"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    headers = [
        "Product",
        "Purchase Price",
        "Selling Price",
        "Current Stock",
        "Minimum Stock",
        "Stock Value",
        "Status"
    ]

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=3,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

    row_num = 4

    for p in products:

        value = (
            float(p.current_stock)
            * float(p.purchase_price)
        )

        status = (
            "Low Stock"
            if float(p.current_stock)
            <= float(p.minimum_stock)
            else "In Stock"
        )

        ws.cell(row=row_num,column=1,value=p.product_name)
        ws.cell(row=row_num,column=2,value=float(p.purchase_price))
        ws.cell(row=row_num,column=3,value=float(p.selling_price))
        ws.cell(row=row_num,column=4,value=float(p.current_stock))
        ws.cell(row=row_num,column=5,value=float(p.minimum_stock))
        ws.cell(row=row_num,column=6,value=value)
        ws.cell(row=row_num,column=7,value=status)

        row_num += 1

    ws.cell(
        row=row_num,
        column=5,
        value="TOTAL VALUE"
    ).font = Font(
        bold=True
    )

    ws.cell(
        row=row_num,
        column=6,
        value=inventory_value
    ).font = Font(
        bold=True
    )

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 15

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=inventory_report.xlsx"
        }
    )

# ----------------------------------------------------
# Booking Report
# ----------------------------------------------------


@router.get(
    "/reports/booking",
    name="booking_report",
)
async def booking_report(
    request: Request,
    from_date: str = "",
    to_date: str = "",
    status: str = "",
    payment_mode: str = "",
    booking_source: str = "",
    user=Depends(login_required),
    db: Session = Depends(get_db),
):
    from datetime import datetime as dt

    from app.booking.models import Booking
    from app.booking.models import BookingRoom
    from app.booking.models import Room
    from app.booking.models import RoomType

    query = (
        db.query(Booking)
        .order_by(
            Booking.check_in_at.desc(),
            Booking.id.desc(),
        )
    )

    normalized_from_date = from_date.strip()
    normalized_to_date = to_date.strip()
    normalized_status = status.strip().upper()
    normalized_payment_mode = payment_mode.strip().upper()
    normalized_booking_source = booking_source.strip().upper()

    if normalized_from_date:
        try:
            parsed_from_date = dt.strptime(
                normalized_from_date,
                "%Y-%m-%d",
            )
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Invalid from_date",
            )

        query = query.filter(
            Booking.check_in_at >= parsed_from_date
        )

    if normalized_to_date:
        try:
            parsed_to_date = dt.strptime(
                normalized_to_date,
                "%Y-%m-%d",
            )
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Invalid to_date",
            )

        parsed_to_date = parsed_to_date.replace(
            hour=23,
            minute=59,
            second=59,
            microsecond=999999,
        )

        query = query.filter(
            Booking.check_in_at <= parsed_to_date
        )

    if normalized_status:
        query = query.filter(
            Booking.status == normalized_status
        )

    if normalized_payment_mode:
        query = query.filter(
            Booking.payment_mode
            == normalized_payment_mode
        )

    if normalized_booking_source:
        query = query.filter(
            Booking.booking_source
            == normalized_booking_source
        )

    bookings = query.all()

    booking_ids = [
        booking.id
        for booking in bookings
    ]

    assignment_rows = []

    if booking_ids:
        assignment_rows = (
            db.query(
                BookingRoom,
                Room,
                RoomType,
            )
            .join(
                Room,
                Room.id == BookingRoom.room_id,
            )
            .join(
                RoomType,
                RoomType.id == Room.room_type_id,
            )
            .filter(
                BookingRoom.booking_id.in_(booking_ids)
            )
            .order_by(
                BookingRoom.booking_id,
                Room.room_number,
            )
            .all()
        )

    assignments_by_booking = {}

    for assignment, room, room_type in assignment_rows:
        assignments_by_booking.setdefault(
            assignment.booking_id,
            [],
        ).append(
            {
                "room_number": room.room_number,
                "room_type": room_type.name,
                "status": assignment.status,
                "cancelled_at": assignment.cancelled_at,
                "cancellation_reason": (
                    assignment.cancellation_reason
                ),
            }
        )

    rows = []

    for booking in bookings:
        assignments = assignments_by_booking.get(
            booking.id,
            [],
        )

        active_rooms = [
            item["room_number"]
            for item in assignments
            if item["status"] == "ACTIVE"
        ]

        cancelled_rooms = [
            item["room_number"]
            for item in assignments
            if item["status"] == "CANCELLED"
        ]

        rows.append(
            {
                "booking": booking,
                "active_rooms": active_rooms,
                "cancelled_rooms": cancelled_rooms,
                "assignments": assignments,
            }
        )

    total_bookings = len(bookings)

    confirmed_bookings = sum(
        1
        for booking in bookings
        if booking.status == "CONFIRMED"
    )

    cancelled_bookings = sum(
        1
        for booking in bookings
        if booking.status == "CANCELLED"
    )

    financial_bookings = [
        booking
        for booking in bookings
        if booking.status != "CANCELLED"
    ]

    gross_booking_value = sum(
        float(booking.total_amount or 0)
        for booking in financial_bookings
    )

    advance_collected = sum(
        float(booking.advance_amount or 0)
        for booking in financial_bookings
    )

    balance_pending = sum(
        max(
            float(booking.total_amount or 0)
            - float(booking.advance_amount or 0),
            0,
        )
        for booking in financial_bookings
    )

    payment_modes = sorted(
        {
            str(booking.payment_mode).strip().upper()
            for booking in (
                db.query(Booking)
                .filter(
                    Booking.payment_mode.isnot(None)
                )
                .all()
            )
            if str(booking.payment_mode).strip()
        }
    )

    return templates.TemplateResponse(
        request=request,
        name="reports/booking.html",
        context={
            "user": user,
            "rows": rows,
            "total_bookings": total_bookings,
            "confirmed_bookings": confirmed_bookings,
            "cancelled_bookings": cancelled_bookings,
            "gross_booking_value": gross_booking_value,
            "advance_collected": advance_collected,
            "balance_pending": balance_pending,
            "from_date": normalized_from_date,
            "to_date": normalized_to_date,
            "status": normalized_status,
            "payment_mode": normalized_payment_mode,
            "payment_modes": payment_modes,
            "booking_source": normalized_booking_source,
            "booking_sources": ("ERP", "ONLINE"),
        },
    )
